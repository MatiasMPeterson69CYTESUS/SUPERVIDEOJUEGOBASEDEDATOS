# timesplit_game.py — TimeSplit (Dragoncito Edition) + PostgreSQL (SQLAlchemy) + Glicko-2
# --------------------------------------------------------------------------------------
# - Juego simple (Carreras y Fútbol) que registra "splits" (marcas) en fracciones de tiempo (ms)
# - Perso principal + bots + Dragoncito con sprite assets/dragon.png (si existe)
# - Guarda sesiones, splits, jugadores, matches y ratings Glicko-2 en PostgreSQL
# - Lee DATABASE_URL desde .env (python-dotenv). Si no existe, cae a SQLite local.
#
# Controles:
#   Menú: ↑/↓ navega · ENTER elegir · 1..6 personaje · M mute · ESC salir
#   Juego: ENTER nueva · ESPACIO pausa · R reiniciar · TAB cambia modo · L vuelta/periodo
#          [ y ] tick (50–1000 ms) · - y + duración (modo)
#   Carreras: ↑/↓ velocidad
#   Fútbol: Flechas moverse · F chutar
#   Guardado: S guardar (DB) · E CSV · X Excel (sesión) · U sync API (opcional)
#
# Requisitos (requirements.txt recomendado):
#   pygame-ce
#   SQLAlchemy
#   psycopg2-binary
#   python-dotenv
#   requests
#   openpyxl
#
# Nota: Si instalas con Python 3.13 en Windows, usa pygame-ce (evita compilación local).

from __future__ import annotations

import os
import csv
import time
import uuid
import math
import random
import re
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple

# --- cargar .env (si existe) ---
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# --- pygame ---
import pygame as pg

# --- requests (opcional) ---
try:
    import requests
except Exception:
    requests = None  # type: ignore

# --- excel export (opcional) ---
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None  # type: ignore

# --- SQLAlchemy ---
from sqlalchemy import (
    create_engine, Column, Integer, Float, String, Text, Boolean,
    ForeignKey, UniqueConstraint, Index, select
)
from sqlalchemy.orm import declarative_base, sessionmaker, relationship

# ======================================================================================
# DB / ORM
# ======================================================================================

Base = declarative_base()

def _db_url() -> str:
    # Prioridad: DATABASE_URL (ideal para Postgres). Si no, SQLite local.
    url = (os.getenv("DATABASE_URL") or "").strip()
    if url:
        return url
    return "sqlite:///timesplit.sqlite"

ENGINE = create_engine(_db_url(), echo=False, future=True)
SessionLocal = sessionmaker(bind=ENGINE, autoflush=False, autocommit=False, future=True)

class OrganizationORM(Base):
    __tablename__ = "organizations"
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String, unique=True, nullable=False)

    players = relationship("PlayerORM", back_populates="organization")

class PlayerORM(Base):
    __tablename__ = "players"
    id = Column(Integer, primary_key=True, autoincrement=True)
    org_id = Column(Integer, ForeignKey("organizations.id", ondelete="CASCADE"), nullable=False, index=True)
    name = Column(String, nullable=False, index=True)
    is_bot = Column(Boolean, default=False, nullable=False)

    # Glicko-2 params almacenados en escala "rating" (tipo Elo), RD y volatility
    rating = Column(Float, default=1500.0, nullable=False)  # r
    rd = Column(Float, default=350.0, nullable=False)       # RD
    vol = Column(Float, default=0.06, nullable=False)       # sigma

    organization = relationship("OrganizationORM", back_populates="players")

    __table_args__ = (UniqueConstraint("org_id", "name", name="uix_org_playername"),)

class GameSessionORM(Base):
    __tablename__ = "game_sessions"
    id = Column(String, primary_key=True)                 # uid
    org_id = Column(Integer, ForeignKey("organizations.id", ondelete="CASCADE"), nullable=False, index=True)
    player_name = Column(String, nullable=False)
    mode = Column(String, nullable=False)                 # "carreras" | "futbol"
    started_at = Column(Integer, nullable=False)          # epoch ms
    duration_ms = Column(Integer, nullable=False)
    total_score = Column(Float, nullable=False)

    splits = relationship("SplitORM", cascade="all, delete-orphan", back_populates="session")

class SplitORM(Base):
    __tablename__ = "splits"
    id = Column(Integer, primary_key=True, autoincrement=True)
    session_id = Column(String, ForeignKey("game_sessions.id", ondelete="CASCADE"), index=True)
    t_ms = Column(Integer, nullable=False)
    lap = Column(Integer, nullable=False)
    score = Column(Float, nullable=False)
    note = Column(Text, nullable=True)

    session = relationship("GameSessionORM", back_populates="splits")
    __table_args__ = (Index("ix_splits_session_t", "session_id", "t_ms"),)

class MatchORM(Base):
    __tablename__ = "matches"
    id = Column(String, primary_key=True)                 # uid
    org_id = Column(Integer, ForeignKey("organizations.id", ondelete="CASCADE"), nullable=False, index=True)
    played_at = Column(Integer, nullable=False)           # epoch ms

    mode = Column(String, nullable=False)
    session_id = Column(String, ForeignKey("game_sessions.id", ondelete="SET NULL"), nullable=True)

    p1_id = Column(Integer, ForeignKey("players.id", ondelete="CASCADE"), nullable=False, index=True)
    p2_id = Column(Integer, ForeignKey("players.id", ondelete="CASCADE"), nullable=False, index=True)

    score1 = Column(Float, nullable=False)
    score2 = Column(Float, nullable=False)

    # outcome1: 1 win, 0 loss, 0.5 draw (para p1)
    outcome1 = Column(Float, nullable=False)

    __table_args__ = (Index("ix_matches_players_time", "p1_id", "p2_id", "played_at"),)

def orm_init_db() -> None:
    Base.metadata.create_all(ENGINE)

def orm_get_or_create_org(name: str = "TimeSplit League") -> int:
    with SessionLocal() as db:
        org = db.execute(select(OrganizationORM).where(OrganizationORM.name == name)).scalar_one_or_none()
        if org:
            return org.id
        org = OrganizationORM(name=name)
        db.add(org)
        db.commit()
        db.refresh(org)
        return org.id

def orm_get_or_create_player(org_id: int, name: str, is_bot: bool) -> PlayerORM:
    with SessionLocal() as db:
        p = db.execute(
            select(PlayerORM).where(PlayerORM.org_id == org_id, PlayerORM.name == name)
        ).scalar_one_or_none()
        if p:
            return p
        p = PlayerORM(org_id=org_id, name=name, is_bot=is_bot, rating=1500.0, rd=350.0, vol=0.06)
        db.add(p)
        db.commit()
        db.refresh(p)
        return p

def orm_save_session_with_splits(org_id: int, payload: Dict) -> None:
    """
    Inserta/actualiza GameSession por id y reemplaza splits.
    """
    with SessionLocal() as db:
        sid = payload["id"]
        ses = db.get(GameSessionORM, sid)
        if not ses:
            ses = GameSessionORM(
                id=sid,
                org_id=org_id,
                player_name=payload["player"],
                mode=payload["mode"],
                started_at=int(payload["startedAt"]),
                duration_ms=int(payload["durationMs"]),
                total_score=float(payload["totalScore"]),
            )
            db.add(ses)
        else:
            ses.player_name = payload["player"]
            ses.mode = payload["mode"]
            ses.started_at = int(payload["startedAt"])
            ses.duration_ms = int(payload["durationMs"])
            ses.total_score = float(payload["totalScore"])
            ses.splits.clear()

        for sp in payload.get("splits", []):
            ses.splits.append(SplitORM(
                session_id=sid,
                t_ms=int(sp["t"]),
                lap=int(sp["lap"]),
                score=float(sp["score"]),
                note=sp.get("note"),
            ))
        db.commit()

def orm_save_match(org_id: int, session_id: str, mode: str, p1: PlayerORM, p2: PlayerORM,
                   score1: float, score2: float, outcome1: float) -> str:
    mid = f"m_{uuid.uuid4().hex[:10]}"
    with SessionLocal() as db:
        m = MatchORM(
            id=mid,
            org_id=org_id,
            played_at=int(time.time() * 1000),
            mode=mode,
            session_id=session_id,
            p1_id=p1.id,
            p2_id=p2.id,
            score1=float(score1),
            score2=float(score2),
            outcome1=float(outcome1),
        )
        db.add(m)
        db.commit()
    return mid

def orm_update_player_glicko(pid: int, rating: float, rd: float, vol: float) -> None:
    with SessionLocal() as db:
        p = db.get(PlayerORM, pid)
        if not p:
            return
        p.rating = float(rating)
        p.rd = float(rd)
        p.vol = float(vol)
        db.commit()

def orm_leaderboard_glicko(org_id: int, limit: int = 10) -> List[Dict]:
    with SessionLocal() as db:
        rows = db.execute(
            select(PlayerORM.name, PlayerORM.rating, PlayerORM.rd, PlayerORM.vol)
            .where(PlayerORM.org_id == org_id)
            .order_by(PlayerORM.rating.desc())
            .limit(limit)
        ).all()
        return [{"player": r[0], "rating": float(r[1]), "rd": float(r[2]), "vol": float(r[3])} for r in rows]

# ======================================================================================
# Glicko-2
# ======================================================================================

GLICKO_SCALE = 173.7178

def _g(phi: float) -> float:
    return 1.0 / math.sqrt(1.0 + 3.0 * (phi ** 2) / (math.pi ** 2))

def _E(mu: float, mu_j: float, phi_j: float) -> float:
    return 1.0 / (1.0 + math.exp(-_g(phi_j) * (mu - mu_j)))

def glicko2_update(
    r: float, RD: float, sigma: float,
    opps: List[Tuple[float, float, float]],
    tau: float = 0.5
) -> Tuple[float, float, float]:
    mu = (r - 1500.0) / GLICKO_SCALE
    phi = RD / GLICKO_SCALE

    if not opps:
        phi_star = math.sqrt(phi * phi + sigma * sigma)
        return r, phi_star * GLICKO_SCALE, sigma

    opp_mus = [((rj - 1500.0) / GLICKO_SCALE, RDj / GLICKO_SCALE, sj) for (rj, RDj, sj) in opps]

    v_inv = 0.0
    delta_sum = 0.0
    for mu_j, phi_j, s_j in opp_mus:
        E_ = _E(mu, mu_j, phi_j)
        g_ = _g(phi_j)
        v_inv += (g_ ** 2) * E_ * (1.0 - E_)
        delta_sum += g_ * (s_j - E_)
    v = 1.0 / v_inv
    delta = v * delta_sum

    a = math.log(sigma * sigma)

    def f(x: float) -> float:
        ex = math.exp(x)
        num = ex * (delta * delta - phi * phi - v - ex)
        den = 2.0 * (phi * phi + v + ex) ** 2
        return (num / den) - ((x - a) / (tau * tau))

    A = a
    if delta * delta > (phi * phi + v):
        B = math.log(delta * delta - phi * phi - v)
    else:
        k = 1
        B = a - k * tau
        while f(B) < 0:
            k += 1
            B = a - k * tau

    fA = f(A)
    fB = f(B)
    for _ in range(60):
        C = A + (A - B) * fA / (fB - fA)
        fC = f(C)
        if abs(fC) < 1e-10:
            A = C
            break
        if fC * fB < 0:
            A = B
            fA = fB
        else:
            fA = fA / 2.0
        B = C
        fB = fC

    sigma_prime = math.exp(A / 2.0)

    phi_star = math.sqrt(phi * phi + sigma_prime * sigma_prime)
    phi_prime = 1.0 / math.sqrt((1.0 / (phi_star * phi_star)) + (1.0 / v))
    mu_prime = mu + (phi_prime * phi_prime) * delta_sum

    r_prime = 1500.0 + mu_prime * GLICKO_SCALE
    RD_prime = phi_prime * GLICKO_SCALE
    return r_prime, RD_prime, sigma_prime

def apply_single_match_glicko(p1: PlayerORM, p2: PlayerORM, outcome1: float, tau: float = 0.5):
    r1, rd1, s1 = p1.rating, p1.rd, outcome1
    r2, rd2, s2 = p2.rating, p2.rd, (1.0 - outcome1) if outcome1 in (0.0, 1.0) else 0.5
    nr1, nrd1, nvol1 = glicko2_update(r1, rd1, p1.vol, [(r2, rd2, s1)], tau=tau)
    nr2, nrd2, nvol2 = glicko2_update(r2, rd2, p2.vol, [(r1, rd1, s2)], tau=tau)
    return (nr1, nrd1, nvol1), (nr2, nrd2, nvol2)

# ======================================================================================
# Juego
# ======================================================================================

WIDTH, HEIGHT = 960, 540
FPS = 60
ASSETS_DIR = "assets"

def uid(prefix="s") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:10]}"

def fmt_ms(ms: int) -> str:
    s = ms // 1000
    mm = s // 60
    ss = s % 60
    cs = (ms % 1000) // 10
    return f"{mm:02d}:{ss:02d}.{cs:02d}"

def safe_filename(text: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]+', "_", text)
    return name.strip().strip(".")

def load_img(name: str):
    path = os.path.join(ASSETS_DIR, name)
    if not os.path.exists(path):
        return None
    try:
        return pg.image.load(path).convert_alpha()
    except Exception:
        return None

def load_snd(name: str, volume: float = 0.6):
    if not hasattr(pg, "mixer"):
        return None
    for candidate in (name, os.path.splitext(name)[0] + ".wav", os.path.splitext(name)[0] + ".ogg"):
        path = os.path.join(ASSETS_DIR, candidate)
        if os.path.exists(path):
            try:
                snd = pg.mixer.Sound(path)
                snd.set_volume(volume)
                return snd
            except Exception:
                continue
    return None

CHARACTERS = [
    {"name": "Aqua",       "color": ( 60,200,255)},
    {"name": "Lime",       "color": ( 80,220,120)},
    {"name": "Rose",       "color": (235, 80,140)},
    {"name": "Gold",       "color": (245,200, 40)},
    {"name": "Violet",     "color": (170, 95,255)},
    {"name": "Dragoncito", "color": (100,200,100)},
]
BOT_NAMES_RACE = ["Bot-Alpha", "Bot-Bravo", "Bot-Charlie", "Bot-Delta", "Bot-Echo"]
BOT_NAMES_FOOT = ["Rival-1", "Rival-2", "Rival-3", "Compi-1", "Compi-2"]

PU_TURBO  = "TURBO"
PU_SHIELD = "ESCUDO"
PU_FIRE   = "FIREBALL"
PU_FREEZE = "FREEZE"
PU_DRAGON = "DRAGON"

@dataclass
class Split:
    t: int
    score: float
    lap: int
    note: Optional[str] = None

class Session:
    def __init__(self, player: str, mode: str):
        self.id = uid("s")
        self.player = player or "Jugador/a"
        self.mode = mode
        self.startedAt = int(time.time() * 1000)
        self.totalScore = 0.0
        self.durationMs = 0
        self.splits: List[Split] = []

class Game:
    def __init__(self):
        orm_init_db()
        self.org_id = orm_get_or_create_org("TimeSplit League")

        pg.init()
        pg.display.set_caption("TimeSplit — Dragoncito Edition (PostgreSQL + Glicko-2)")
        self.screen = pg.display.set_mode((WIDTH, HEIGHT))
        self.clock = pg.time.Clock()
        self.font = pg.font.SysFont("consolas,arial", 18)
        self.big = pg.font.SysFont("consolas,arial", 28, bold=True)

        self.muted = False
        try:
            pg.mixer.init()
        except Exception:
            pass
        self.snd_pick  = load_snd("s_pick.ogg", 0.7)  or load_snd("s_pick.wav", 0.7)
        self.snd_shoot = load_snd("s_shoot.ogg", 0.7) or load_snd("s_shoot.wav", 0.7)
        self.snd_goal  = load_snd("s_goal.ogg", 0.8)  or load_snd("s_goal.wav", 0.8)

        self.dragon_img = load_img("dragon.png")

        self.screen_state = "menu"
        self.menu_idx = 0

        self.player_name = os.getenv("TSR_PLAYER") or "Jugador/a"
        self.player_character_idx = 5
        self.mode = "carreras"
        self.tick_ms = 200
        self.target_duration_s = 60
        self.half_duration_s = 45
        self.api_url = os.getenv("TSR_API") or "http://localhost:3000/api/sessions"

        self.running = False
        self.paused = False
        self.elapsed_ms = 0
        self.score = 0.0
        self.lap = 1
        self.enemy_score = 0
        self.last_tick = 0

        self.session: Optional[Session] = None
        self.last_saved_payload: Optional[dict] = None
        self.message = ""
        self.msg_until = 0

        self.speed = 20.0
        self.cars: List[dict] = []

        self.player_pos = pg.Vector2(120, HEIGHT / 2)
        self.ball_pos = pg.Vector2(WIDTH / 2, HEIGHT / 2)
        self.ball_vel = pg.Vector2(0, 0)
        self.npcs: List[dict] = []

        self.powerups: List[dict] = []
        self.active_pu: Dict[str, int] = {}
        self.next_pu_spawn_ms = 2500

        for bn in BOT_NAMES_RACE + BOT_NAMES_FOOT:
            orm_get_or_create_player(self.org_id, bn, is_bot=True)

    def info(self, text: str, ms: int = 2200):
        self.message = text
        self.msg_until = pg.time.get_ticks() + ms

    def play_snd(self, snd):
        if self.muted or snd is None:
            return
        try:
            snd.play()
        except Exception:
            pass

    def _init_race_bots(self):
        lanes = [HEIGHT * 0.30, HEIGHT * 0.38, HEIGHT * 0.46, HEIGHT * 0.54, HEIGHT * 0.62]
        random.shuffle(lanes)
        self.cars = []

        ch = CHARACTERS[self.player_character_idx % len(CHARACTERS)]
        self.player_label = ch["name"]
        self.player_color = ch["color"]

        self.cars.append({
            "name": f"{self.player_label} ({self.player_name})",
            "color": self.player_color,
            "x": 60, "y": lanes[0],
            "speed": self.speed,
            "dist": 0.0,
            "is_player": True
        })

        for i in range(1, min(5, len(lanes))):
            self.cars.append({
                "name": BOT_NAMES_RACE[i - 1],
                "color": (180, 180, 200) if i % 2 == 0 else (200, 150, 80),
                "x": random.randint(20, 180),
                "y": lanes[i],
                "speed": random.uniform(12, 28),
                "dist": 0.0,
                "is_player": False
            })

    def _init_football_npcs(self):
        self.npcs = []
        for i in range(3):
            self.npcs.append({
                "name": BOT_NAMES_FOOT[i],
                "color": (210, 80, 80),
                "pos": pg.Vector2(random.randint(WIDTH // 2 + 40, WIDTH - 60), random.randint(60, HEIGHT - 60)),
                "role": "rival"
            })
        for i in range(2):
            self.npcs.append({
                "name": BOT_NAMES_FOOT[3 + i],
                "color": (80, 180, 250),
                "pos": pg.Vector2(random.randint(60, WIDTH // 2 - 60), random.randint(60, HEIGHT - 60)),
                "role": "ally"
            })

    def get_limit_ms(self) -> int:
        if self.mode == "carreras":
            return int(self.target_duration_s * 1000)
        return int(2 * self.half_duration_s * 1000)

    def start_session(self):
        self.session = Session(self.player_name, self.mode)
        self.running = True
        self.paused = False
        self.elapsed_ms = 0
        self.score = 0.0
        self.lap = 1
        self.enemy_score = 0
        self.last_tick = 0

        self.powerups.clear()
        self.active_pu.clear()
        self.next_pu_spawn_ms = 2500

        self.player_pos.update(120, HEIGHT / 2)
        self.ball_pos.update(WIDTH / 2, HEIGHT / 2)
        self.ball_vel.update(0, 0)

        self._init_race_bots()
        self._init_football_npcs()

        self.info("Sesión iniciada (ENTER nueva, ESPACIO pausa)")
        orm_get_or_create_player(self.org_id, self.player_name, is_bot=False)

    def _register_split_tick(self):
        if not self.session:
            return
        self.session.splits.append(Split(t=int(self.elapsed_ms), score=round(self.score, 2), lap=int(self.lap)))

    def register_event(self, note: str):
        if not self.session:
            return
        self.session.splits.append(Split(t=int(self.elapsed_ms), score=round(self.score, 2), lap=int(self.lap), note=note))

    def finish_session(self):
        if not self.session:
            return
        self.running = False
        limit = self.get_limit_ms()
        self.session.totalScore = round(self.score, 2)
        self.session.durationMs = max(self.elapsed_ms, limit)

        payload = {
            "id": self.session.id,
            "player": self.session.player,
            "mode": self.session.mode,
            "startedAt": self.session.startedAt,
            "durationMs": self.session.durationMs,
            "totalScore": self.session.totalScore,
            "splits": [{"t": s.t, "lap": s.lap, "score": s.score, "note": s.note} for s in self.session.splits],
        }

        orm_save_session_with_splits(self.org_id, payload)
        self.last_saved_payload = payload
        self._persist_match_and_update_glicko()
        self.info("Guardado OK (DB) + Rating actualizado (Glicko-2)")

    def _persist_match_and_update_glicko(self):
        p1 = orm_get_or_create_player(self.org_id, self.player_name, is_bot=False)

        if self.mode == "carreras":
            bot = None
            best_dist = -1.0
            player_dist = 0.0
            for c in self.cars:
                if c.get("is_player"):
                    player_dist = float(c.get("dist", 0.0))
                else:
                    if float(c.get("dist", 0.0)) > best_dist:
                        best_dist = float(c.get("dist", 0.0))
                        bot = c
            if bot is None:
                bot_name = BOT_NAMES_RACE[0]
                bot_score = 0.0
            else:
                bot_name = bot["name"]
                bot_score = best_dist
            score1 = player_dist
            score2 = bot_score
            outcome1 = 0.5 if abs(score1 - score2) < 1e-9 else (1.0 if score1 > score2 else 0.0)
        else:
            bot_name = "Rival-1"
            score1 = float(self.score)
            score2 = float(self.enemy_score)
            outcome1 = 0.5 if score1 == score2 else (1.0 if score1 > score2 else 0.0)

        p2 = orm_get_or_create_player(self.org_id, bot_name, is_bot=True)

        orm_save_match(self.org_id, self.session.id, self.mode, p1, p2, score1, score2, outcome1)
        (nr1, nrd1, nvol1), (nr2, nrd2, nvol2) = apply_single_match_glicko(p1, p2, outcome1, tau=0.5)
        orm_update_player_glicko(p1.id, nr1, nrd1, nvol1)
        orm_update_player_glicko(p2.id, nr2, nrd2, nvol2)

    def export_csv_last(self):
        if not self.last_saved_payload:
            self.info("No hay sesión guardada aún")
            return
        p = self.last_saved_payload
        fname = f"timesplit_{safe_filename(p['mode'])}_{safe_filename(p['player'])}_{p['startedAt']}.csv"
        header = ["player", "mode", "startedAt", "durationMs", "totalScore", "t(ms)", "lap/periodo", "score", "note"]
        try:
            with open(fname, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(header)
                for sp in p.get("splits", []):
                    w.writerow([p["player"], p["mode"], p["startedAt"], p["durationMs"], p["totalScore"],
                                sp["t"], sp["lap"], sp["score"], sp.get("note") or ""])
            self.info(f"CSV exportado: {fname}")
        except Exception as e:
            self.info(f"Error exportando CSV: {e}")

    def export_xlsx_last(self):
        if openpyxl is None:
            self.info("Falta openpyxl (pip install openpyxl)")
            return
        if not self.last_saved_payload:
            self.info("No hay sesión guardada aún")
            return
        p = self.last_saved_payload
        fname = f"timesplit_{safe_filename(p['mode'])}_{safe_filename(p['player'])}_{p['startedAt']}.xlsx"
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "splits"
            ws.append(["player", "mode", "startedAt", "durationMs", "totalScore"])
            ws.append([p["player"], p["mode"], p["startedAt"], p["durationMs"], p["totalScore"]])
            ws.append([])
            ws.append(["t(ms)", "lap/periodo", "score", "note"])
            for sp in p.get("splits", []):
                ws.append([sp["t"], sp["lap"], sp["score"], sp.get("note") or ""])
            for col in range(1, 5):
                ws.column_dimensions[get_column_letter(col)].width = 18
            ws.column_dimensions[get_column_letter(4)].width = 28
            wb.save(fname)
            self.info(f"Excel exportado: {fname}")
        except Exception as e:
            self.info(f"Error exportando Excel: {e}")

    def _spawn_powerup(self):
        kinds = [PU_TURBO, PU_SHIELD, PU_FIRE, PU_FREEZE, PU_DRAGON]
        self.powerups.append({"kind": random.choice(kinds),
                              "pos": pg.Vector2(random.randint(120, WIDTH - 120), random.randint(90, HEIGHT - 90)),
                              "t": self.elapsed_ms})

    def _update_powerups(self):
        if self.elapsed_ms >= self.next_pu_spawn_ms:
            self._spawn_powerup()
            self.next_pu_spawn_ms += random.randint(2200, 4200)
        pickup_radius = 22
        player_pos = self.player_pos if self.mode == "futbol" else pg.Vector2(140, self.cars[0]["y"] if self.cars else HEIGHT/2)
        new_list = []
        for pu in self.powerups:
            if pu["pos"].distance_to(player_pos) <= pickup_radius:
                self.play_snd(self.snd_pick)
                self.active_pu[pu["kind"]] = self.elapsed_ms + 3500
                self.register_event(f"PICK {pu['kind']}")
            else:
                new_list.append(pu)
        self.powerups = new_list
        for k in list(self.active_pu.keys()):
            if self.elapsed_ms >= self.active_pu[k]:
                del self.active_pu[k]

    def _update_carreras(self, dt_ms: int):
        speed = self.speed * (1.35 if PU_TURBO in self.active_pu else 1.0)
        if self.cars:
            self.cars[0]["speed"] = speed
        for c in self.cars:
            c["dist"] += float(c["speed"]) * (dt_ms / 1000.0)
            c["x"] = 60 + (c["dist"] * 10) % (WIDTH - 140)
        if self.cars:
            self.score = float(self.cars[0]["dist"])

    def _update_futbol(self, dt_ms: int):
        self.ball_pos += self.ball_vel * (dt_ms / 16.0)
        self.ball_vel *= 0.992
        if self.ball_pos.y < 18 or self.ball_pos.y > HEIGHT - 18:
            self.ball_vel.y *= -1
        if self.ball_pos.x < 18 or self.ball_pos.x > WIDTH - 18:
            self.ball_vel.x *= -1

        for npc in self.npcs:
            to_ball = self.ball_pos - npc["pos"]
            if to_ball.length() > 1:
                to_ball.scale_to_length(0.7 if npc["role"] == "rival" else 0.45)
                npc["pos"] += to_ball

        goal_top, goal_bot = int(HEIGHT * 0.35), int(HEIGHT * 0.65)
        if self.ball_pos.x <= 10 and goal_top <= self.ball_pos.y <= goal_bot:
            self.enemy_score += 1
            self.play_snd(self.snd_goal)
            self.register_event("GOAL EN CONTRA")
            self.ball_pos.update(WIDTH / 2, HEIGHT / 2)
            self.ball_vel.update(0, 0)
        if self.ball_pos.x >= WIDTH - 10 and goal_top <= self.ball_pos.y <= goal_bot:
            self.score += 1
            self.play_snd(self.snd_goal)
            self.register_event("GOAL A FAVOR")
            self.ball_pos.update(WIDTH / 2, HEIGHT / 2)
            self.ball_vel.update(0, 0)

    def _draw_hud(self):
        pg.draw.rect(self.screen, (15, 18, 30), (0, 0, WIDTH, 58))
        t = self.font.render(f"Modo: {self.mode} | Tick: {self.tick_ms}ms | Duración objetivo: {self.get_limit_ms()//1000}s", True, (220, 230, 255))
        self.screen.blit(t, (14, 10))
        t2 = self.big.render(f"{fmt_ms(self.elapsed_ms)}", True, (255, 255, 255))
        self.screen.blit(t2, (14, 28))
        if self.mode == "carreras":
            s = self.big.render(f"Puntaje(dist): {self.score:.2f} | Vuelta: {self.lap}", True, (200, 255, 200))
        else:
            s = self.big.render(f"Goles: {int(self.score)} (Rivales {self.enemy_score}) | Periodo: {self.lap}", True, (200, 255, 200))
        self.screen.blit(s, (250, 26))
        if self.active_pu:
            active = " ".join(list(self.active_pu.keys()))
            ptxt = self.font.render(f"Power-ups: {active}", True, (255, 210, 120))
            self.screen.blit(ptxt, (14, 56))
        if self.message and pg.time.get_ticks() < self.msg_until:
            msg = self.font.render(self.message, True, (255, 220, 220))
            self.screen.blit(msg, (14, HEIGHT - 24))

    def _draw_carreras(self):
        pg.draw.rect(self.screen, (30, 30, 35), (0, int(HEIGHT * 0.22), WIDTH, int(HEIGHT * 0.56)))
        pg.draw.line(self.screen, (240, 240, 240), (0, HEIGHT // 2), (WIDTH, HEIGHT // 2), 3)
        pg.draw.rect(self.screen, (220, 40, 80), (WIDTH - 18, int(HEIGHT * 0.22), 18, int(HEIGHT * 0.56)))
        for c in self.cars:
            x = int(c["x"]); y = int(c["y"])
            if c.get("is_player") and self.player_label == "Dragoncito" and self.dragon_img is not None:
                img = pg.transform.smoothscale(self.dragon_img, (52, 52))
                self.screen.blit(img, (x, y - 26))
            else:
                pg.draw.rect(self.screen, c["color"], (x, y - 18, 44, 36), border_radius=6)
            name = self.font.render(c["name"], True, (220, 220, 220))
            self.screen.blit(name, (x + 50, y - 10))
        for pu in self.powerups:
            pg.draw.circle(self.screen, (255, 210, 120), (int(pu["pos"].x), int(pu["pos"].y)), 10)

    def _draw_futbol(self):
        self.screen.fill((10, 120, 70))
        pg.draw.rect(self.screen, (255, 255, 255), (12, 76, WIDTH - 24, HEIGHT - 120), 2)
        pg.draw.line(self.screen, (255, 255, 255), (WIDTH // 2, 76), (WIDTH // 2, HEIGHT - 44), 2)
        pg.draw.circle(self.screen, (255, 255, 255), (WIDTH // 2, HEIGHT // 2 + 16), 42, 2)
        goal_top, goal_bot = int(HEIGHT * 0.35), int(HEIGHT * 0.65)
        pg.draw.rect(self.screen, (255, 255, 255), (0, goal_top, 12, goal_bot - goal_top), 2)
        pg.draw.rect(self.screen, (255, 255, 255), (WIDTH - 12, goal_top, 12, goal_bot - goal_top), 2)
        ch = CHARACTERS[self.player_character_idx % len(CHARACTERS)]
        if ch["name"] == "Dragoncito" and self.dragon_img is not None:
            img = pg.transform.smoothscale(self.dragon_img, (52, 52))
            self.screen.blit(img, (int(self.player_pos.x - 26), int(self.player_pos.y - 26)))
        else:
            pg.draw.circle(self.screen, ch["color"], (int(self.player_pos.x), int(self.player_pos.y)), 12)
        for npc in self.npcs:
            pg.draw.circle(self.screen, npc["color"], (int(npc["pos"].x), int(npc["pos"].y)), 10)
        pg.draw.circle(self.screen, (250, 250, 250), (int(self.ball_pos.x), int(self.ball_pos.y)), 7)
        for pu in self.powerups:
            pg.draw.circle(self.screen, (255, 210, 120), (int(pu["pos"].x), int(pu["pos"].y)), 10)

    def _handle_menu_event(self, ev):
        if ev.type != pg.KEYDOWN:
            return
        if ev.key == pg.K_ESCAPE:
            raise SystemExit
        if ev.key == pg.K_DOWN:
            self.menu_idx = (self.menu_idx + 1) % 4
        if ev.key == pg.K_UP:
            self.menu_idx = (self.menu_idx - 1) % 4
        if ev.key == pg.K_m:
            self.muted = not self.muted
            self.info("Mute ON" if self.muted else "Mute OFF")
        if pg.K_1 <= ev.key <= pg.K_6:
            self.player_character_idx = int(ev.unicode) - 1
            self.info(f"Personaje: {CHARACTERS[self.player_character_idx]['name']}")
        if ev.key == pg.K_RETURN:
            if self.menu_idx == 0:
                self.mode = "carreras"; self.start_session(); self.screen_state = "game"
            elif self.menu_idx == 1:
                self.mode = "futbol"; self.start_session(); self.screen_state = "game"
            elif self.menu_idx == 2:
                self.screen_state = "ranking"
            else:
                raise SystemExit

    def _handle_game_event(self, ev):
        if ev.type != pg.KEYDOWN:
            return
        if ev.key == pg.K_ESCAPE:
            self.screen_state = "menu"; self.running = False; return
        if ev.key == pg.K_SPACE:
            self.paused = not self.paused; self.info("Pausa" if self.paused else "Reanudar")
        if ev.key == pg.K_RETURN:
            self.start_session()
        if ev.key == pg.K_TAB:
            self.mode = "futbol" if self.mode == "carreras" else "carreras"
            self.start_session(); self.info(f"Modo: {self.mode}")
        if ev.key == pg.K_r:
            self.start_session()
        if ev.key == pg.K_l:
            self.lap = min(self.lap + 1, 99)
            self.register_event("LAP/PERIODO")
        if ev.key == pg.K_LEFTBRACKET:
            self.tick_ms = max(50, self.tick_ms - 50); self.info(f"Tick: {self.tick_ms} ms")
        if ev.key == pg.K_RIGHTBRACKET:
            self.tick_ms = min(1000, self.tick_ms + 50); self.info(f"Tick: {self.tick_ms} ms")
        if ev.key == pg.K_MINUS:
            if self.mode == "carreras": self.target_duration_s = max(10, self.target_duration_s - 5)
            else: self.half_duration_s = max(15, self.half_duration_s - 5)
        if ev.key in (pg.K_PLUS, pg.K_EQUALS):
            if self.mode == "carreras": self.target_duration_s = min(180, self.target_duration_s + 5)
            else: self.half_duration_s = min(120, self.half_duration_s + 5)
        if ev.key == pg.K_s:
            self.finish_session()
        if ev.key == pg.K_e:
            self.export_csv_last()
        if ev.key == pg.K_x:
            self.export_xlsx_last()
        if self.mode == "carreras":
            if ev.key == pg.K_UP: self.speed = min(200.0, self.speed + 2.0)
            if ev.key == pg.K_DOWN: self.speed = max(0.0, self.speed - 2.0)
        else:
            if ev.key == pg.K_f: self.shoot()

    def shoot(self):
        d = self.ball_pos - self.player_pos
        dist = d.length()
        if 0 < dist < 42:
            power = 7.5 + random.random() * 5.0
            d.scale_to_length(power * 20)
            self.ball_vel = d
            self.play_snd(self.snd_shoot)
            self.register_event("SHOT")

    def _draw_menu(self):
        self.screen.fill((8, 10, 22))
        title = self.big.render("TimeSplit — Dragoncito (Postgres + Glicko-2)", True, (240, 240, 255))
        self.screen.blit(title, (40, 40))
        subtitle = self.font.render("↑/↓ elegir · ENTER · 1..6 personaje · M mute", True, (190, 200, 230))
        self.screen.blit(subtitle, (40, 78))
        opts = ["Jugar Carreras", "Jugar Fútbol", "Ver Ranking (Glicko-2)", "Salir"]
        y = 140
        for i, o in enumerate(opts):
            col = (255, 255, 255) if i == self.menu_idx else (170, 180, 210)
            self.screen.blit(self.big.render(o, True, col), (70, y))
            y += 54
        ch = CHARACTERS[self.player_character_idx % len(CHARACTERS)]
        p = self.font.render(f"Jugador: {self.player_name} | Personaje: {ch['name']} | DB: {'Postgres' if os.getenv('DATABASE_URL') else 'SQLite'}", True, (210, 210, 210))
        self.screen.blit(p, (40, HEIGHT - 36))

    def _draw_ranking(self):
        self.screen.fill((10, 10, 18))
        self.screen.blit(self.big.render("Ranking (Glicko-2) — TOP 10", True, (240, 240, 255)), (40, 36))
        self.screen.blit(self.font.render("ESC volver al menú", True, (170, 180, 210)), (40, 66))
        rows = orm_leaderboard_glicko(self.org_id, limit=10)
        y = 120
        self.screen.blit(self.font.render("Pos   Jugador                Rating    RD     Vol", True, (220, 220, 220)), (60, y))
        y += 22
        for idx, r in enumerate(rows, start=1):
            line = f"{idx:>2}   {r['player'][:20]:<20}   {r['rating']:>7.1f}  {r['rd']:>6.1f}  {r['vol']:.4f}"
            col = (200, 230, 200) if idx == 1 else (200, 200, 200)
            self.screen.blit(self.font.render(line, True, col), (60, y))
            y += 22

    def run(self):
        while True:
            dt = self.clock.tick(FPS)
            for ev in pg.event.get():
                if ev.type == pg.QUIT:
                    raise SystemExit
                if self.screen_state == "menu":
                    self._handle_menu_event(ev)
                elif self.screen_state == "game":
                    self._handle_game_event(ev)
                elif self.screen_state == "ranking":
                    if ev.type == pg.KEYDOWN and ev.key == pg.K_ESCAPE:
                        self.screen_state = "menu"

            if self.screen_state == "game" and self.running and not self.paused:
                self.elapsed_ms += dt
                self._update_powerups()
                if self.mode == "carreras":
                    self._update_carreras(dt)
                else:
                    keys = pg.key.get_pressed()
                    v = pg.Vector2(0, 0)
                    if keys[pg.K_UP]: v.y -= 1
                    if keys[pg.K_DOWN]: v.y += 1
                    if keys[pg.K_LEFT]: v.x -= 1
                    if keys[pg.K_RIGHT]: v.x += 1
                    if v.length() > 0:
                        v = v.normalize() * (3.4 if PU_TURBO in self.active_pu else 2.6)
                        self.player_pos += v
                        self.player_pos.x = max(30, min(WIDTH - 30, self.player_pos.x))
                        self.player_pos.y = max(90, min(HEIGHT - 60, self.player_pos.y))
                    self._update_futbol(dt)

                if self.elapsed_ms - self.last_tick >= self.tick_ms:
                    self.last_tick = self.elapsed_ms
                    self._register_split_tick()

                if self.elapsed_ms >= self.get_limit_ms():
                    self.finish_session()

            if self.screen_state == "menu":
                self._draw_menu()
            elif self.screen_state == "ranking":
                self._draw_ranking()
            else:
                if self.mode == "carreras":
                    self.screen.fill((8, 10, 22))
                    self._draw_carreras()
                else:
                    self._draw_futbol()
                self._draw_hud()

            pg.display.flip()

if __name__ == "__main__":
    Game().run()
